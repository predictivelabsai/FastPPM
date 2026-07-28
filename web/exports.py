"""CSV / XLSX export of extracted and canonical tables.

Tables are built server-side from the data (the document's extracted entities or
an initiative's canonical milestones/risks), then returned as a download. XLSX is
styled with the FastPPM header; CSV is plain. Mirrors the sister PEHero export route.
"""

from __future__ import annotations

import csv
import io

from starlette.responses import Response

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── Table builders ──────────────────────────────────────────────────────────

def _doc_table(data: dict, kind: str):
    """(title, columns, rows) from a document's extracted structured_json."""
    if kind == "risks":
        rows = [[r.get("description"), r.get("probability"), r.get("impact"),
                 r.get("mitigation")] for r in data.get("risks", [])]
        return "Risks", ["Risk", "Probability", "Impact", "Mitigation"], rows
    if kind == "value_drivers":
        rows = [[v.get("name"), v.get("category"), v.get("target"), v.get("realized"),
                 v.get("period")] for v in data.get("value_drivers", [])]
        return "Value drivers", ["Driver", "Category", "Target", "Realised", "Period"], rows
    rows = [[m.get("title"), m.get("owner"), m.get("planned_date"), m.get("actual_date"),
             m.get("progress"), m.get("status")] for m in data.get("milestones", [])]
    return "Milestones", ["Title", "Owner", "Planned date", "Actual date",
                          "Progress %", "Status"], rows


def initiative_table(store, iid: int, kind: str):
    """(title, columns, rows) from an initiative's canonical data."""
    if kind == "risks":
        rows = [[r.get("description"), r.get("probability"), r.get("impact"),
                 r.get("mitigation"), r.get("status")] for r in store.list_risks(iid)]
        return "Risks", ["Risk", "Probability", "Impact", "Mitigation", "Status"], rows
    rows = [[m.get("title"), m.get("owner"), m.get("baseline_date"), m.get("planned_date"),
             m.get("actual_date"), m.get("progress"), m.get("status")]
            for m in store.list_milestones(iid)]
    return "Milestones", ["Milestone", "Owner", "Baseline date", "Planned date",
                          "Actual date", "Progress %", "Status"], rows


def document_table(store, did: int, kind: str):
    d = store.get_document(did)
    if not d:
        return None
    title, cols, rows = _doc_table(d.get("structured_json") or {}, kind)
    return d["file_name"], title, cols, rows


# ── Responses ───────────────────────────────────────────────────────────────

def _clean(v):
    return "" if v is None else v


def csv_response(columns, rows, filename: str) -> Response:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(columns)
    for r in rows:
        w.writerow([_clean(c) for c in r])
    return Response(buf.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def xlsx_response(title, columns, rows, filename: str) -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Export")[:31]
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="6B1766", end_color="6B1766", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="E6E3EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, border
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=_clean(val))
            c.border = border
    for ci, col in enumerate(columns, 1):
        width = len(str(col))
        for row in rows[:30]:
            width = max(width, len(str(_clean(row[ci - 1])) if ci - 1 < len(row) else ""))
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(width + 4, 48)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.read(), media_type=XLSX_MEDIA,
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})
